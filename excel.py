"""Excel input/output layer for the ETL pipeline."""

from openpyxl import Workbook, load_workbook

from config import EXCEL_PATH
from parser import process_phone


async def process_excel(client) -> None:
    """Read inputs, enrich rows, and persist results after each item."""
    if not EXCEL_PATH.exists():
        EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        demo_wb = Workbook()
        demo_ws = demo_wb.active
        demo_ws.append(["phone"])
        demo_ws.append(["+380XXXXXXXXX"])
        demo_ws.append(["+380YYYYYYYYY"])
        demo_wb.save(EXCEL_PATH)
        print(f"Created demo input: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if headers[:1] != ["phone"]:
        raise ValueError("The first column must be named 'phone'.")

    # Portfolio version uses safe, non-sensitive output fields.
    output_headers = ["phone", "status", "source"]
    for column, header in enumerate(output_headers, start=1):
        ws.cell(row=1, column=column, value=header)

    total = max(ws.max_row - 1, 0)

    for index, row in enumerate(range(2, ws.max_row + 1), start=1):
        phone = ws.cell(row=row, column=1).value
        if phone is None or str(phone).strip() == "":
            continue

        phone = str(phone).strip()
        print("=" * 60)
        print(f"[{index}/{total}] Input: {phone}")

        try:
            data = await process_phone(client, phone)
            ws.cell(row=row, column=2, value=data["status"])
            ws.cell(row=row, column=3, value=data["source"])
            print(f"Status: {data['status']}")
        except Exception as exc:
            ws.cell(row=row, column=2, value="error")
            ws.cell(row=row, column=3, value=type(exc).__name__)
            print(f"Error: {type(exc).__name__}")

        wb.save(EXCEL_PATH)

    wb.save(EXCEL_PATH)
    print("\nPipeline completed.")
